from _catch22 import Catch22
import pycatch22
import numpy as np
from aeon.datasets import tsc_datasets, load_italy_power_demand, load_arrow_head, load_basic_motions, load_acsf1
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from decimal import Decimal, ROUND_HALF_UP
import csv
import os

feature_names_short = [
    "mode_5",
    "mode_10",
    "acf_timescale",
    "acf_first_min",
    "ami2",
    "trev",
    "high_fluctuation",
    "stretch_high",
    "transition_matrix",
    "periodicity",
    "embedding_dist",
    "ami_timescale",
    "whiten_timescale",
    "outlier_timing_pos",
    "outlier_timing_neg",
    "centroid_freq",
    "stretch_decreasing",
    "entropy_pairs",
    "rs_range",
    "dfa",
    "low_freq_power",
    "forecast_error"
]

features_names_pycatch22 = [
        'DN_HistogramMode_5',
        'DN_HistogramMode_10',
        'CO_f1ecac',
        'CO_FirstMin_ac',
        'CO_HistogramAMI_even_2_5',
        'CO_trev_1_num',
        'MD_hrv_classic_pnn40',
        'SB_BinaryStats_mean_longstretch1',
        'SB_TransitionMatrix_3ac_sumdiagcov',
        'PD_PeriodicityWang_th0_01',
        'CO_Embed2_Dist_tau_d_expfit_meandiff',
        'IN_AutoMutualInfoStats_40_gaussian_fmmi',
        'FC_LocalSimple_mean1_tauresrat',
        'DN_OutlierInclude_p_001_mdrmd',
        'DN_OutlierInclude_n_001_mdrmd',
        'SP_Summaries_welch_rect_area_5_1',
        'SB_BinaryStats_diff_longstretch0',
        'SB_MotifThree_quantile_hh',
        'SC_FluctAnal_2_rsrangefit_50_1_logi_prop_r1',
        'SC_FluctAnal_2_dfa_50_1_2_logi_prop_r1',
        'SP_Summaries_welch_rect_centroid',
        'FC_LocalSimple_mean3_stderr'
]

# $env:NUMBA_DISABLE_JIT = "0"
# echo $env:NUMBA_DISABLE_JIT
# Numba Disabled Switcher 0 = off, 1 = on

IPD_X_train, IPD_y_train = load_basic_motions(split="train")
IPD_X_train = [IPD_X_train[0]]
test = IPD_X_train[0][0][0:20]
IPD_X_train = np.array([[test]])
print("Training Data: ", IPD_X_train)
print("Length of data: ", len(IPD_X_train[0][0]))
os.environ['NUMBA_DISABLE_JIT'] = '0'
print("Numba Off: ",os.environ['NUMBA_DISABLE_JIT'])


#aeon
aeon_c22 = Catch22(replace_nans=True)
_ = aeon_c22.fit_transform(IPD_X_train)

results_aeon = [features_names_pycatch22]
for i in range(len(_)):
    #formatting it to pycatch22 format
    results_aeon.append(_[i])

#aeon.pycatch22
aeon_pycatch_c22 = Catch22(use_pycatch22=True, replace_nans=True)
pycResult = aeon_pycatch_c22.fit_transform(IPD_X_train)

results_pycatch22 = [features_names_pycatch22]
for i in range(len(pycResult)):
    #formatting it to pycatch22 format
    results_pycatch22.append(pycResult[i])


# #pycatch
# results_pycatch22 = [features_names_pycatch22]
# # for i in range(len(IPD_X_train)):
# results = pycatch22.catch22_all(IPD_X_train[i][0])
# results_pycatch22.append(results['values'])





"Workbook"
aeon_file_name = "aeon_catch22_with_numba"

#aeon xlsx
wb = Workbook()
ws = wb.active
for i in range(len(results_aeon)):
    for j in range(len(results_aeon[i])):
        cell = ws.cell(row=i + 1, column=j + 1, value=results_aeon[i][j])
        if(i != 0):
            rounded_aeon_results = Decimal(results_aeon[i][j]).quantize(Decimal('.0000'), 
                                                                    rounding=ROUND_HALF_UP)
            rounded_pycatch22_results = Decimal(results_pycatch22[i][j]).quantize(Decimal('.0000'), 
                                                                                rounding=ROUND_HALF_UP)
            if rounded_aeon_results != rounded_pycatch22_results:
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

wb.save(aeon_file_name + ".xlsx")

#pycatch xlsx
wb = Workbook()
ws = wb.active
for i in range(len(results_pycatch22)):
    for j in range(len(results_pycatch22[i])):
        cell = ws.cell(row=i + 1, column=j + 1, value=results_pycatch22[i][j])
wb.save("pycatch22_catch22_ipd.xlsx")


#aeon.csv
with open(aeon_file_name + ".csv", mode='w', newline='') as file:
    writer = csv.writer(file)
    for data in results_aeon:
        writer.writerow(data)

#pycatch.csv  
with open('pycatch22_catch22_ipd.csv', mode='w', newline='') as file:
    writer = csv.writer(file)
    for data in results_pycatch22:
        writer.writerow(data)

print("Finished writing data")